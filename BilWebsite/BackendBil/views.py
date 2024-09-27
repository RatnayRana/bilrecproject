from django.shortcuts import render
import re
import pdfplumber
import pandas as pd
import os
from django.db import IntegrityError
import logging
from decimal import Decimal
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from .models import BankStatement
from .models import Bank, System, DailyReportBankStatement, BankAccount
from django.db.models import Q
from datetime import date, datetime
from django.db import connection
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.styles import NamedStyle
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_date
import openpyxl

pattern_lms = r"\b\d{9}\b"

pattern_pf = r"^(\d{5,})\s*-\s*(.*)"
pattern_insurance = r"^\d{9}+$"
rr_pattern = re.compile(r":(\d+)/")
record_pattern = re.compile(r"\d{2}-[A-Z]{3}-\d{2}")
description_pattern = re.compile(r"(?<=\d{2}-[A-Z]{3}-\d{2})(?!.*\b\d{3}[A-Z]+\d{10,}|\d{16}\b).*")
pattern_reference = re.compile(r"\b\d{3}[A-Za-z]+\d+|\d{16}\b")
instrument_pattern = re.compile(r"(?<![-/#,&\d])\b\d{6}\b(?![-/#,&\d]| [A-Z])")
pattern_amount = re.compile(r"\d{1,3}(?:,\d{3})*\.\d{2}")
cheque_cleared = re.compile(r"(#|NO|[.])\s*(\d{6})")
in_house = re.compile(r"(?:-)\s*(\d{6})")
Incoming_fund_transfer = re.compile(r"(?:RRN-|RRN\s*-)\s*(\d+)")
bank_number_regex = re.compile(r"STATEMENT OF ACCOUNT FOR\s*[:]*\s*(\d{6,})")
Bob_Bank_number_regex = re.compile(r"ACCOUNT NO.\s*[:]*\s*(\d{6,})")
date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")
logger = logging.getLogger(__name__)


def clean_cell(cell):
    if cell is not None:
        return cell.replace("\n", " ").strip()
    return cell


cursor = connection.cursor()


def index(request):

    return render(request, "loginpage.html")
    # return render(request, 'home.html')


from django.contrib.auth.decorators import login_required
import logging

logger = logging.getLogger(__name__)


@login_required
def home(request):
    user = request.user

    # Access user details
    username = user.username
    print(username)
    logger.info(f"Logged in user: {username}")  #
    return render(request, "home.html")


def user(request):

    return render(request, "userpage.html")


def uploadBankStatement(request):
    user = request.user

    # Access user details
    user_id = user.id

    folder = "my_folder/"

    if request.method == "POST":
        system_name = request.POST.getlist("system_name[]")
        bank_name = request.POST.getlist("bank_name[]")
        uploaded_file = request.FILES.getlist("file[]")

        for bank_name, system_name, uploaded_file in zip(
            bank_name, system_name, uploaded_file
        ):
            Bank_data = None
            System_data = None
            user_uploaded = None

            try:
                Bank_data = Bank.objects.get(Q(name=bank_name))
                System_data = System.objects.get(Q(name=system_name))
                user_uploaded = User.objects.get(id=user_id)
                print(System_data)

            except System.DoesNotExist:
                print("theerer")
                return render(
                    request,
                    "uploaddailyreport.html",
                    {"error": "No such system present"},
                )
            except Bank.DoesNotExist:
                print("theerer")
                return render(
                    request,
                    "uploaddailyreport.html",
                    {"error": "No such bank name present"},
                )
            except User.DoesNotExist:
                return render(
                    request,
                    "uploaddailyreport.html",
                    {"error": "No such user  present"},
                )

            except:
                messages.error(request, "No data  present")
                return render(request, "uploadbankstatement.html")

                # uploaded_file = request.FILES.get('file') or request.FILES.get('file_drop')
                # print(Bank_data)

            if uploaded_file:
                fs = FileSystemStorage(location=folder)
                filename = fs.save(uploaded_file.name, uploaded_file)

                file_path = fs.path(filename)
            else:

                messages.error(request, "Please Upload the file please")
                return render(
                    request,
                    "uploadbankstatement.html",
                    {"error": "Please upload the file"},
                )

            if bank_name == "Bank of Bhutan":
                Bob_Bank_number = None

                if os.path.exists(file_path):

                    pdf = pdfplumber.open(
                        file_path
                    )  # Use the absolute path to open the file
                    data_save = False
                    clean_data = []
                    for page in pdf.pages:
                        raw_text = page.extract_text()
                        if raw_text:
                            # Split text by lines and then combine lines based on record boundaries
                            lines = raw_text.split("\n")
                            for Line in lines:
                                bank_number = Bob_Bank_number_regex.search(Line)
                                if bank_number is not None:
                                    Bob_Bank_number = bank_number.group(1)

                        count_table = page.extract_table()
                        try:
                            cleaned_table = [
                                [clean_cell(cell) for cell in row]
                                for row in count_table
                            ]
                            df = pd.DataFrame(
                                cleaned_table[1:], columns=cleaned_table[0]
                            )
                            df = df.drop("VALUE DATE", axis=1)

                            for index, row in df.iterrows():
                                particulars = row.get("PARTICULARS", "")
                                if particulars and not particulars.endswith("88888"):
                                    match_data = rr_pattern.search(particulars)
                                    if match_data:
                                        df.at[index, "PARTICULARS"] = match_data.group(
                                            1
                                        )
                                    else:
                                        df.at[index, "PARTICULARS"] = ""

                            else:
                                df.at[index, "PARTICULARS"] = ""
                            data_dicts = df.to_dict(orient="records")
                            clean_data.extend(data_dicts)
                        except:

                            messages.error(request, "You have uploaded the wrong file")
                            break

                    for record in clean_data:
                        try:
                            date = datetime.strptime(record["POST DATE"], "%d/%m/%Y")
                        except ValueError:
                            print(
                                f"Skipping record with invalid date: {record['POST DATE']}"
                            )
                            continue
                        journal_number = (
                            record.get("journal_number", "").strip() or None
                        )
                        balance = (
                            float(record["BALANCE"].replace(",", ""))
                            if record["BALANCE"]
                            else None
                        )
                        balance = round(balance, 2) if balance is not None else None

                        # if not BankStatement.objects.filter(
                        # date=date,
                        # journal_number=journal_number).exists():
                        try:

                            BankStatement.objects.create(
                                date=date,
                                journal_number=record["JOURNAL NUMBER"],
                                rr_number=(
                                    record["PARTICULARS"]
                                    if record["PARTICULARS"]
                                    else None
                                ),
                                debit=(
                                    float(record["DEBIT"].replace(",", ""))
                                    if record["DEBIT"]
                                    else None
                                ),
                                credit=(
                                    float(record["CREDIT"].replace(",", ""))
                                    if record["CREDIT"]
                                    else None
                                ),
                                instrument_number=(
                                    record["CHEQUE NO/ REFERENCE"]
                                    if record["CHEQUE NO/ REFERENCE"]
                                    else None
                                ),
                                balance=balance,
                                bank_name=Bank_data,
                                bank_account_number=Bob_Bank_number,
                                system_name=System_data,
                                useruploaded=user_uploaded,
                            )
                        except IntegrityError:
                            pass
                            # logger.warning(f"Duplicate entry found for date: {date} and journal number: {journal_number}.")

                        data_save = True
                    if data_save:
                        print("The data are saved perfectly")
                        messages.success(request, "The data are saved perfectly")
                    else:
                        messages.error(request, "No valid data to save")
                else:
                    print("Failed to save data")

            elif bank_name == "Bhutan National Bank":
                if os.path.exists(file_path):
                    pdf = pdfplumber.open(file_path)
                    BankStatementSave = False
                    Bank_Number = None
                    combined_lines = []
                    current_record = []
                    formatted_transactions = []
                    for page in pdf.pages:
                        raw_text = page.extract_text()
                        if raw_text:
                            lines = raw_text.split("\n")
                            for line in lines:
                                line = line.strip()

                                # Search for bank number
                                bank_number = bank_number_regex.search(line)
                                if bank_number is not None:
                                    Bank_Number = bank_number.group(1)

                                if record_pattern.match(line):
                                    if current_record:
                                        combined_lines.append(
                                            " ".join(current_record).strip()
                                        )
                                    current_record = [line]
                                else:
                                    current_record.append(line)
                            if current_record:
                                combined_lines.append(" ".join(current_record).strip())
                                current_record = []
                    formatted_records = []
                    for record in combined_lines:
                        if not record_pattern.match(record):
                            if formatted_records:
                                formatted_records[-1] += " " + record
                            else:
                                formatted_records.append(record)
                        else:
                            formatted_records.append(record)

                    for record in formatted_records:
                        print(record)
                        date_match = record_pattern.search(record)
                        date = date_match.group() if date_match else ""
                        description_match = description_pattern.search(record)
                        description = (
                            description_match.group() if description_match else ""
                        )
                        if description.strip().startswith("CHEQUE CLEARED"):
                            cheque_number = cheque_cleared.search(description)
                            if cheque_number:
                                transaction_number = cheque_number.group(2)
                                transaction_status = "CHEQUE CLEARED"

                            else:
                                transaction_number = None
                                transaction_status = None

                        elif description.strip().startswith("In-House"):
                            inhouse_pattern = in_house.search(description)
                            if inhouse_pattern:
                                transaction_number = inhouse_pattern.group(1)
                                transaction_status = "IN-House Cheque"
                            else:
                                transaction_number = None
                                transaction_status = None
                        elif description.strip().startswith("Incoming Fund Transfer"):
                            Incoming = Incoming_fund_transfer.search(description)
                            if Incoming:
                                transaction_number = Incoming.group(1)
                                transaction_status = "INCOMING FUND TRANSFER"
                            else:
                                transaction_number = None
                                transaction_status = None
                        elif description.strip().startswith("Incoming Payment via"):
                            Incoming = Incoming_fund_transfer.search(description)
                            if Incoming:
                                transaction_number = Incoming.group(1)
                                transaction_status = "Incoming Payment via"
                            else:
                                transaction_number = None
                                transaction_status = None
                        else:
                            transaction_number = None
                            transaction_status = None
                        references = pattern_reference.findall(record)
                        reference = references[0] if references else " "
                        instrument_match = instrument_pattern.search(record)
                        instrument = (
                            instrument_match.group() if instrument_match else " "
                        )
                        print(instrument)
                        amounts = pattern_amount.findall(record)
                        credit_balance = amounts[0] if len(amounts) > 0 else ""
                        balance = amounts[1] if len(amounts) > 1 else ""
                        formatted_transactions.append(
                            {
                                "Date": date,
                                "Description": description,
                                "Reference": reference,
                                "CheckNo_InstrumentNo": instrument,
                                "Amount": credit_balance,
                                "Balance": balance,
                                "TransactionNumber": transaction_number,
                                "Transaction_status": transaction_status,
                            }
                        )

                    formatted_transactions = formatted_transactions[2:]
                    for record in formatted_transactions:
                        try:
                            date = datetime.strptime(record["Date"], "%d-%b-%y")
                            BankStatement.objects.create(
                                date=date,
                                journal_number=(
                                    record["TransactionNumber"]
                                    if record["TransactionNumber"]
                                    else None
                                ),
                                # rr_number= record['Description'] if record['Description'] else None,
                                instrument_number=(
                                    record["CheckNo_InstrumentNo"]
                                    if record["CheckNo_InstrumentNo"]
                                    else None
                                ),
                                reference_no=(
                                    record["Reference"] if record["Reference"] else None
                                ),
                                transaction_type=(
                                    record["Transaction_status"]
                                    if record["Transaction_status"]
                                    else None
                                ),
                                credit=float(record["Amount"].replace(",", "")),
                                balance=float(record["Balance"].replace(",", "")),
                                bank_name=Bank_data,
                                bank_account_number=Bank_Number,
                                system_name=System_data,
                                useruploaded=user_uploaded,
                            )
                            BankStatementSave = True
                        except IntegrityError as e:
                            logger.warning(
                                f"IntegrityError: {e}. Record with date {date} and journal number {journal_number} already exists."
                            )
                        except ValueError as e:
                            print(f"Error parsing date: {e}")

                    if BankStatementSave:
                        print("The data are saved perfectly")
                        messages.success(request, "The data are saved perfectly")
                    else:
                        messages.error(request, "No valid data to save")

            else:
                messages.error(request, "Please Enter the  require details")

    return render(request, "uploadbankstatement.html")


def uploaddailyreport(request):
    user = request.user
    user_id = user.id
    folder = "dailyreport/"
    if request.method == "POST":
        system_names = request.POST.getlist("system_name[]")
        uploaded_files = request.FILES.getlist("file[]")
        print(system_names, uploaded_files)
        if len(system_names) != len(uploaded_files):
            messages.error(request, "Mismatch between system names and uploaded files")
            return render(
                request,
                "uploaddailyreport.html",
                {"error": "Mismatch between system names and uploaded files"},
            )

        for n, system_name in enumerate(system_names):
            uploaded_file = uploaded_files[n]
            try:
                System_data = System.objects.get(Q(name=system_name))
                user_uploaded = User.objects.get(id=user_id)
                print(f"Processing {system_name}")

            except System.DoesNotExist:
                messages.error(request, f"System '{system_name}' does not exist")
                continue
            except User.DoesNotExist:
                messages.error(request, "User does not exist")
                return render(
                    request,
                    "uploaddailyreport.html",
                    {"error": "No such User is present"},
                )
            except Exception as e:
                print(e)
                messages.error(request, f"Error retrieving data: {e}")
                continue

            if uploaded_file:
                fs = FileSystemStorage(location=folder)
                filename = fs.save(uploaded_file.name, uploaded_file)
                file_path = fs.path(filename)
            else:
                messages.error(request, f"No file uploaded for {system_name}")
                continue

            try:
                if system_name in ["PF", "GF"]:
                    process_pf_gf(file_path, System_data, user_uploaded)
                elif system_name == "Insurance":
                    process_insurance(file_path, System_data, user_uploaded)
                elif system_name == "LMS":
                    process_lms(file_path, System_data, user_uploaded)
                else:
                    messages.warning(request, f"Unknown system: {system_name}")
                    continue

                messages.success(request, f"Data for {system_name} saved successfully")
            except Exception as e:
                messages.error(request, f"Error processing {system_name}: {e}")

        return render(request, "uploaddailyreport.html")

    return render(request, "uploaddailyreport.html")


def process_pf_gf(file_path, System_data, user_uploaded):
    if os.path.exists(file_path):
        account_number = None
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        first_sheet = wb.worksheets[0]
        for row in first_sheet.iter_rows():
            for cell in row:
                match = re.search(pattern_pf, str(cell.value))
                if match:  # Check if the regex pattern matches
                    account_number = match.group(1)
                    print(f"Account Number: {account_number}")
        clean_data = []
        for row in ws.iter_rows():
            for cell in row:
                for cell in row:
                    if not cell.border.left.style:
                        try:
                            cell.value = None
                        except:
                            continue
        df = (
            pd.DataFrame(ws.values)
            .dropna(how="all", axis=0)
            .dropna(how="all", axis=1)
            .reset_index(drop=True)
        )  # load into pandas, drop empty rows and columns
        df = df.T.set_index(0).T
        date_format = "%Y-%m-%d"
        df["Instrument Date"] = pd.to_datetime(
            df["Instrument Date"], format=date_format, errors="coerce"
        ).dt.date
        df["Instrument Date"] = df["Instrument Date"].apply(
            lambda x: x if pd.notna(x) else None
        )

        data_dicts = df.to_dict(orient="records")

        try:
            for record in data_dicts:
                if not DailyReportBankStatement.objects.filter(
                    voucher_no=record["Voucher No"]
                ).exists():

                    DailyReportBankStatement.objects.create(
                        tran_date=(
                            record["Trans.Date"].strftime("%Y-%m-%d")
                            if record["Trans.Date"]
                            else None
                        ),
                        voucher_no=(
                            record["Voucher No"] if record["Voucher No"] else None
                        ),
                        instrument_number=(
                            record["Instrument No"] if record["Instrument No"] else None
                        ),
                        instrument_date=(
                            record["Instrument Date"]
                            if record["Instrument Date"]
                            else None
                        ),
                        credit_amount=(
                            float(record["Cr Amount"]) if record["Cr Amount"] else 0
                        ),
                        debit_amount=(
                            float(record["Dr Amount"]) if record["Dr Amount"] else 0
                        ),
                        system_name=System_data,
                        bank_account_number=account_number,
                        useruploaded=user_uploaded,
                    )
                BankStatement = True
        except Exception as e:
            print(f"There was a problem uploading the data: {e}")


def process_insurance(file_path, System_data, user_uploaded):
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        first_sheet = wb.worksheets[0]
        account_number = None
        for row in first_sheet.iter_rows():
            for cell in row:
                if not cell.border.left.style:
                    match = re.search(pattern_insurance, str(cell.value))
                    if match:  # Check if the regex pattern matches
                        account_number = match.group()
                        # print(f"Account Number: {account_number}")
        clean_data = []
        for row in ws.iter_rows():
            for cell in row:
                for cell in row:
                    if not cell.border.left.style:
                        try:
                            cell.value = None
                        except:
                            continue
        df = (
            pd.DataFrame(ws.values)
            .dropna(how="all", axis=0)
            .dropna(how="all", axis=1)
            .reset_index(drop=True)
        )  # load into pandas, drop empty rows and columns
        df = df.T.set_index(0).T
        date_format = "%Y-%m-%d"
        df["Instrument Date"] = pd.to_datetime(
            df["Instrument Date"], format=date_format, errors="coerce"
        ).dt.date
        df["Instrument Date"] = df["Instrument Date"].apply(
            lambda x: x if pd.notna(x) else None
        )

        data_dicts = df.to_dict(orient="records")

        try:
            for record in data_dicts:
                date_str = record.get("Date")
                if date_str:
                    try:
                        date_obj = datetime.strptime(
                            date_str, "%d-%b-%Y"
                        )  # Adjust format if needed
                        formatted_date = date_obj.strftime("%Y-%m-%d")
                    except ValueError:
                        formatted_date = None
                else:
                    formatted_date = None
                if not DailyReportBankStatement.objects.filter(
                    voucher_no=record["Voucher No"]
                ).exists():

                    DailyReportBankStatement.objects.create(
                        tran_date=formatted_date,
                        voucher_no=(
                            record["Voucher No"] if record["Voucher No"] else None
                        ),
                        instrument_number=(
                            record["Instrument No"] if record["Instrument No"] else None
                        ),
                        instrument_date=(
                            record["Instrument Date"]
                            if record["Instrument Date"]
                            else None
                        ),
                        credit_amount=(
                            float(record["Cr Amount"]) if record["Cr Amount"] else 0
                        ),
                        debit_amount=(
                            float(record["Dr Amount"]) if record["Dr Amount"] else 0
                        ),
                        system_name=System_data,
                        bank_account_number=account_number,
                        useruploaded=user_uploaded,
                    )
                BankStatement = True
        except Exception as e:
            print(f"There was a problem uploading the data: {e}")


def process_lms(file_path, System_data, user_uploaded):
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        first_sheet = wb.worksheets[0]
        print(first_sheet)
        account_number = None
        excel_data = pd.read_excel(file_path, sheet_name=None)
        if 'Sheet1' in excel_data:
            sheet_data = excel_data["Sheet1"]
            print(sheet_data)
        else:
    # If Sheet1 doesn't exist, handle the error, perhaps by using the first sheet
            sheet_data = list(excel_data.values())[0]  # Select the first available sheet
            print("Sheet1 not found. Using the first available sheet.")
            print(sheet_data)

        combined_text = " ".join(
            sheet_data.iloc[2:5]
            .fillna("")
            .apply(lambda row: " ".join(row.values.astype(str)), axis=1)
        )
        matches = re.search(pattern_lms, combined_text)
        if matches:
            account_number = matches.group()
            print(account_number)
        clean_data = []
        for row in ws.iter_rows():
            for cell in row:
                for cell in row:
                    if not cell.border.left.style:
                        try:
                            cell.value = None
                        except:
                            continue
        df = (
            pd.DataFrame(ws.values)
            .dropna(how="all", axis=0)
            .dropna(how="all", axis=1)
            .reset_index(drop=True)
        )  # load into pandas, drop empty rows and columns
        df = df.T.set_index(0).T

        data_dicts = df.to_dict(orient="records")
        # print(data_dicts)

        try:
            for record in data_dicts:
                if not DailyReportBankStatement.objects.filter(
                    voucher_no=record["Voucher No"]
                ).exists():

                    DailyReportBankStatement.objects.create(
                        tran_date=(
                            record["Voucher Date"].strftime("%Y-%d-%m")
                            if record["Voucher Date"]
                            else None
                        ),
                        voucher_no=(
                            record["Voucher No"] if record["Voucher No"] else None
                        ),
                        instrument_number=(
                            record["Cheque/ Account No"]
                            if record["Cheque/ Account No"]
                            else None
                        ),
                        credit_amount=(
                            float(record["Cr Amount"]) if record["Cr Amount"] else 0
                        ),
                        debit_amount=(
                            float(record["Dr Amount"]) if record["Dr Amount"] else 0
                        ),
                        system_name=System_data,
                        bank_account_number=account_number,
                        useruploaded=user_uploaded,
                    )
                BankStatement = True
        except Exception as e:
            print(f"There was a problem uploading the data: {e}")


def generateReport(request):
    context = {}
    account_numbers = BankAccount.objects.values_list("account_number", "account_name")
    system = System.objects.values_list("id", "name")
    #  AND dr.system_name_id = %s
    #  AND bbs.bank_account_number = %s
    # system_id,account_number

    # AND dr.system_name_id = %s
    # AND bbs.bank_account_number= %s
    context = {"account_numbers": account_numbers, "system": system}
    if request.method == "POST":
        system_id = None
        startdate = request.POST.get("start_date")
        enddate = request.POST.get("end_date")
        account_number = request.POST.get("account_number")
        if not (startdate and enddate and account_number):
            messages.error(request, "Please select all the required fields please")
            return render(request, "generateReport.html", context)

        systemname = request.POST.get("typeofsystem")

        if systemname:
            try:
                system = System.objects.get(name=systemname)
                print(system)
                system_id = system.id
                print(system_id)
            except System.DoesNotExist:
                # Handle the case where the systemname does not exist in the database
                print(f"System with name '{systemname}' does not exist.")
        else:
            # No systemname provided, system_id remains None
            print("No system name provided.")

        try:

            if startdate > enddate:
                messages.error(request, "Start date cannot be after end date.")
                return render(request, "generateReport.html", context)

        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
            return render(request, "generateReport.html", context)

        wb = Workbook()

        # Remove the default sheet
        wb.remove(wb.active)

        # Create sheets
        sheet1 = wb.create_sheet("Reconciliation Report")
        sheet2 = wb.create_sheet("FailedBank statement")
        sheet3 = wb.create_sheet("Failed reportSOS")

        try:
            query = """
            SELECT DISTINCT ON (dr.instrument_number)
                    dr.id,
                    dr.tran_date,
                    dr.voucher_no,
                    dr.instrument_number,
                    dr.credit_amount,
                    dr.debit_amount,
                    sys.name,
                    bbs.journal_number,
                    bbs.reference_no,
                    bbs.instrument_number AS bbs_instrument_number,
                     bbs.rr_number,
                    bbs.debit,
                    bbs.credit,
                    bbs.bank_account_number,
                    dr.status
                FROM public."BackendBil_dailyreportbankstatement" dr
                LEFT JOIN public."BackendBil_system" sys ON dr.system_name_id = sys.id
                LEFT JOIN public."BackendBil_bankstatement" bbs ON 
                    sys.id = bbs.system_name_id and(
                    dr.instrument_number = bbs.journal_number 
                    OR dr.instrument_number = bbs.instrument_number 
                    OR dr.instrument_number = bbs.reference_no
                    or dr.instrument_number= bbs.rr_number)
 
                WHERE dr.status = 'success' 
                AND dr.tran_date BETWEEN %s AND %s 
                   
                """

        except Exception as e:
            print("your error ", e)
            messages.error(request, "An error occurred while processing your request.")
            return render(request, "generateReport.html", context)
        conditions = []
        params = [startdate, enddate]

        if system_id and system_id != "Select System ID":
            conditions.append("dr.system_name_id = %s")
            params.append(system_id)

        cleaned_account_number = account_number.strip()

        if cleaned_account_number and cleaned_account_number != "Select Account Number":
            conditions.append(
                "bbs.bank_account_number = %s AND  dr.bank_account_number=%s"
            )
            params.append(cleaned_account_number)  # Append the first account number
            params.append(cleaned_account_number)  # Append the first account number
            # print("Constructed Query:", query)
        print("Parameters:", params)

        if conditions:
            query += " AND " + " AND ".join(conditions)
        try:
            try:
                reconciliation_report = DailyReportBankStatement.objects.raw(
                    query, params
                )
                print("heelo")

            except Exception as e:
                print(f"An error occurred: {str(e)}")
                messages.error(
                    request,
                    f"An error occurred while processing your request: {str(e)}",
                )
                return render(request, "generateReport.html", context)
            total_records = len(list(reconciliation_report))

            print(f"Total records retrieved: {total_records}")
            headers = [
                "SL.No",
                "Date",
                "Voucher Number",
                "SOS Instrument Number",
                "SOS credit_amount",
                "SOS debit_amount",
                "System Name",
                "Bank Statement instrument_number",
                "bankstatement_reference_no",
                "bankstatement_journalnumber",
                " bankstatement RR_number" "bankstatement_credit",
                "bankstatement_debit",
                "Bank Number",
                "status",
            ]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                sheet1[f"{col_letter}1"] = header
            for row_num, record in enumerate(reconciliation_report, 2):
                sheet1[f"A{row_num}"] = row_num - 1  # SL.No starts from 1
                sheet1[f"B{row_num}"] = getattr(record, "tran_date", None)
                sheet1[f"C{row_num}"] = getattr(record, "voucher_no", "N/A")
                sheet1[f"D{row_num}"] = getattr(record, "instrument_number", "N/A")
                sheet1[f"E{row_num}"] = getattr(record, "credit_amount", "N/A")
                sheet1[f"F{row_num}"] = getattr(record, "debit_amount", "N/A")
                sheet1[f"G{row_num}"] = getattr(record, "name", "N/A")
                sheet1[f"H{row_num}"] = getattr(record, "instrument_number", "N/A")
                sheet1[f"I{row_num}"] = getattr(record, "reference_no", "N/A")
                sheet1[f"J{row_num}"] = getattr(record, "journal_number", "N/A")
                sheet1[f"K{row_num}"] = getattr(record, "rr_number", "N/A")
                sheet1[f"L{row_num}"] = getattr(record, "credit", "N/A")
                sheet1[f"M{row_num}"] = getattr(record, "debit", "N/A")
                sheet1[f"N{row_num}"] = getattr(record, "bank_account_number", "N/A")
                sheet1[f"O{row_num}"] = getattr(record, "status", "N/A")

        except Exception as e:
            print(f"An error occurred: {str(e)}")
            return HttpResponse(f"An error occurred: {str(e)}", status=500)
        try:
            query = """
                SELECT  
                dr.id, dr.tran_date, dr.voucher_no, dr.instrument_number,
						dr.credit_amount,
						dr.debit_amount, dr.status
                FROM public."BackendBil_dailyreportbankstatement" dr
                WHERE (dr.status = 'Failed')
                AND (dr.tran_date BETWEEN %s AND %s)
                
    
            """
        except Exception as e:
            messages.error(request, "An error occurred while processing your request.")
            return render(request, "generateReport.html", context)

        conditions = []
        params = [startdate, enddate]

        # if system_id and system_id != 'Select System ID':
        #     conditions.append("dr.system_name_id = %s")
        #     params.append(system_id)

        cleaned_account_number = account_number.strip()
        if cleaned_account_number and cleaned_account_number != "Select Account Number":
            conditions.append("dr.bank_account_number=%s")
            params.append(cleaned_account_number)

        if conditions:
            query += " AND " + " AND ".join(conditions)

        print("Constructed Query:", query)
        print("Parameters:", account_number)

        try:
            try:
                reconciliation_report = DailyReportBankStatement.objects.raw(
                    query, params
                )
                print("world")
            except Exception as e:
                print(f"An error occurred: {str(e)}")
                messages.error(
                    request,
                    f"An error occurred while processing your request: {str(e)}",
                )
                return render(request, "generateReport.html", context)
            total_records = len(list(reconciliation_report))
            headers = [
                "SL.No",
                "Date",
                "Voucher Number",
                "SOS Instrument Number",
                "credit_amount",
                "debit_amount",
                "status",
            ]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                sheet3[f"{col_letter}1"] = header
            for row_num, record in enumerate(reconciliation_report, 2):
                sheet3[f"A{row_num}"] = row_num - 1
                sheet3[f"B{row_num}"] = getattr(record, "tran_date", None)
                sheet3[f"C{row_num}"] = getattr(record, "voucher_no", "N/A")
                sheet3[f"D{row_num}"] = getattr(record, "instrument_number", "N/A")
                sheet3[f"E{row_num}"] = getattr(record, "credit_amount", "N/A")
                sheet3[f"F{row_num}"] = getattr(record, "debit_amount", "N/A")

                sheet3[f"G{row_num}"] = getattr(record, "status", "N/A")

        except Exception as e:
            print(f"An error occurred: {str(e)}")
            return HttpResponse(f"An error occurred: {str(e)}", status=500)

        try:
            query = """
                SELECT bbs.id,bbs.date,bbs.journal_number,bbs.instrument_number,bbs.reference_no,bbs.debit,bbs.credit,bbs.bank_account_number,bbs.status FROM public."BackendBil_bankstatement" bbs 
                WHERE bbs.status = 'Failed'
                AND bbs.date BETWEEN %s AND %s
            
         
                
                """
        except Exception as e:
            messages.error(request, "An error occurred while processing your request.")
            return render(request, "generateReport.html", context)
        conditions = []
        params = [startdate, enddate]

        if system_id and system_id != "Select System ID":
            conditions.append("bbs.system_name_id = %s")
            params.append(system_id)

        cleaned_account_number = account_number.strip()

        if cleaned_account_number and cleaned_account_number != "Select Account Number":
            conditions.append("bbs.bank_account_number = %s ")
            params.append(cleaned_account_number)

        if conditions:
            query += " AND " + " AND ".join(conditions)
        try:
            try:
                reconciliation_report = DailyReportBankStatement.objects.raw(
                    query, params
                )
                print("How are you")
            except Exception as e:
                print(f"An error occurred: {str(e)}")
                messages.error(
                    request,
                    f"An error occurred while processing your request: {str(e)}",
                )
                return render(request, "generateReport.html", context)
            total_records = len(list(reconciliation_report))
            headers = [
                "SL.No",
                "Date",
                "Journal Number",
                "Instrument Number",
                "Reference  Number",
                "debit",
                "credit",
                "Bank Account Number",
                "status",
            ]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                sheet2[f"{col_letter}1"] = header
            for row_num, record in enumerate(reconciliation_report, 2):
                sheet2[f"A{row_num}"] = row_num - 1
                sheet2[f"B{row_num}"] = getattr(record, "date", None)
                sheet2[f"C{row_num}"] = getattr(record, "journal_number", "N/A")
                sheet2[f"D{row_num}"] = getattr(record, "instrument_number", "N/A")
                sheet2[f"E{row_num}"] = getattr(record, "reference_no", "N/A")
                sheet2[f"F{row_num}"] = getattr(record, "debit", "N/A")
                sheet2[f"G{row_num}"] = getattr(record, "credit", "N/A")
                sheet2[f"H{row_num}"] = getattr(record, "bank_account_number", "N/A")
                sheet2[f"I{row_num}"] = getattr(record, "status", "N/A")

        except Exception as e:
            print(f"An error occurred: {str(e)}")
            return HttpResponse(f"An error occurred: {str(e)}", status=500)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reconciliation_report_{timestamp}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        # Save the workbook to the response
        wb.save(response)

        return response

    return render(request, "generateReport.html", context)


from django.db import connection


def startReconcilation(request):
    startdate = request.POST.get("start_date")
    enddate = request.POST.get("end_date")
    account_number = request.POST.get("account_number")

    context = {}
    account_numbers = BankAccount.objects.values_list("account_number", "account_name")

    system = System.objects.values_list("id", "name")
    #  AND dr.system_name_id = %s
    #  AND bbs.bank_account_number = %s
    # system_id,account_number

    # AND dr.system_name_id = %s
    # AND bbs.bank_account_number= %s
    context = {"account_numbers": account_numbers, "system": system}

    if startdate and enddate:
        startdate = datetime.strptime(startdate, "%Y-%m-%d").date()
        enddate = datetime.strptime(enddate, "%Y-%m-%d").date()

        success = 0
        Failed = 0

        with connection.cursor() as cursor:
            # Update DailyReportBankStatement based on matching criteria

            cursor.execute(
                """
            UPDATE public."BackendBil_dailyreportbankstatement" drbs
        SET status = CASE
        WHEN EXISTS (
            SELECT 1
            FROM public."BackendBil_bankstatement" bs
            WHERE 
                (
                    bs.journal_number = drbs.instrument_number OR 
                    bs.instrument_number = drbs.instrument_number OR 
                    bs.reference_no = drbs.instrument_number OR
                    bs.rr_number = drbs.instrument_number
                ) 
                AND 
                (
                    bs.debit = drbs.credit_amount OR 
                    bs.credit = drbs.debit_amount
                )
                AND (
                    bs.system_name_id = drbs.system_name_id
                )
                
                AND drbs.tran_date BETWEEN %s AND %s
                AND bs.date BETWEEN %s AND %s
                AND bank_account_number = %s
                
        )
        THEN 'success'
        ELSE 'Failed'
    END
   
""",
                [startdate, enddate, startdate, enddate, account_number],
            )

            cursor.execute(
                """
                UPDATE public."BackendBil_bankstatement" bs
                SET status = CASE
                    WHEN EXISTS (
                        SELECT 1
                        FROM public."BackendBil_dailyreportbankstatement" drbs
                        WHERE 
                            (
                                drbs.instrument_number = bs.journal_number OR 
                                drbs.instrument_number = bs.reference_no OR
                                drbs.instrument_number = bs.instrument_number OR 
                                drbs.instrument_number = bs.rr_number 

                            ) 
                            AND 
                            (
                                drbs.debit_amount = bs.credit OR 
                                drbs.credit_amount = bs.debit
                            )
                           AND(
                            drbs.system_name_id=bs.system_name_id
                           )
                            AND drbs.tran_date BETWEEN %s AND %s
                            AND bs.date BETWEEN %s AND %s
                           AND bank_account_number = %s
                    )
                    THEN 'success'
                    ELSE 'Failed'
                END
              
            """,
                [startdate, enddate, startdate, enddate, account_number],
            )

            # Count the number of successful and failed updates in BankStatement
            cursor.execute(
                """
                SELECT 
                    SUM(CASE WHEN status = 'success' THEN 1 ELSE 0 END) AS success_count,
                    SUM(CASE WHEN status = 'Failed' THEN 1 ELSE 0 END) AS failed_count
                FROM public."BackendBil_bankstatement"
                WHERE date BETWEEN %s AND %s
            """,
                [startdate, enddate],
            )
            result = cursor.fetchone()
            success += result[0] or 0
            Failed += result[1] or 0

            # Display success and failure messages
        if success > 0:
            messages.success(request, f"Reconciliation done for {success} records.")
        if Failed > 0:
            messages.error(request, f"Reconciliation failed for {Failed} records.")

        return render(request, "Reconcialtion.html", context)

    else:

        return render(request, "Reconcialtion.html", context)


from django.views.decorators.http import require_GET
from django.http import JsonResponse


@require_GET
def get_account_numbers(request):
    system_id = request.GET.get("system_id")
    print(system_id)
    system_no = System.objects.get(name=system_id)
    system_main_id = system_no.id
    accounts = BankAccount.objects.filter(system_name_id=system_main_id).values(
        "account_name", "account_number"
    )
    return JsonResponse({"account_numbers": list(accounts)})


from django.contrib.auth.hashers import make_password
from django.contrib.auth import get_user_model

User = get_user_model()


def userpage(request):
    save = False
    if request.method == "POST":
        username = request.POST.get("name")
        employee_id = request.POST.get("empid")
        email = request.POST.get("email")
        cid = request.POST.get("cid")
        password = request.POST.get("password")

        if not all([username, employee_id, email, cid, password]):
            messages.error(request, "All fields are required.")
            print("Missing required fields")
            return render(request, "userpage.html")
        if re.fullmatch(r"[A-Za-z0-9@           #$%^&+=]{8,}", password):
            password = password
            # match
        else:
            print("User with employee ID already exists")
            messages.error(request, "You havent saved the secure password")

        userexist = User.objects.filter(employee_id=employee_id).exists()

        if userexist:
            print("User with employee ID already exists")
            messages.error(request, "A user with this employee ID already exists.")
        else:
            try:
                print(f"Creating user with employee ID: {employee_id}")
                hashpassword = make_password(password)
                User.objects.create(
                    username=username,
                    employee_id=employee_id,
                    email=email,
                    cid=int(cid),  # Convert to integer as cid is BigIntegerField
                    password=hashpassword,
                    status="Active",  # Set an initial status
                )

                messages.success(request, "User created successfully!")
                return redirect("loginpage")
            except Exception as e:
                print(f"Error creating user: {e}")
                messages.error(request, "An error occurred while creating the user.")

    return render(request, "userpage.html")


from django.contrib.auth.hashers import check_password

# from django.shortcuts import render, redirect
# from django.http import HttpResponse


def loginpage(request):
    if request.method == "POST":
        empid = request.POST["empid"]
        password = request.POST["Password"]
        print(password)

        try:
            user = User.objects.get(employee_id=empid)

            if user.password is None:
                return HttpResponse("Password field is empty", status=500)

            print("Stored password hash:", user.password)
            is_password_correct = check_password(password, user.password)
            print("Is password correct?", is_password_correct)

            if is_password_correct:
                login(request, user)
                return redirect("home")
            else:
                # Authentication failed
                return HttpResponse("Invalid credentials", status=401)
        except User.DoesNotExist:
            return HttpResponse("User does not exist", status=404)

    return render(request, "loginpage.html")
